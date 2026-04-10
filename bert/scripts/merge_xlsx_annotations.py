#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PRIMARY_WIDTH = 15
PRIMARY_FILL_MAP = {
    "id": 0,
    "cleaned_text": 1,
    "keyword": 6,
    "broad": 12,
    "strict": 13,
}
NO_HEADER_INDEXES = {
    "id": 0,
    "cleaned_text": 1,
    "keyword": 6,
    "broad": 12,
    "strict": 13,
}


def is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return normalize_value(value).lower()


def comparable_value(value: Any, field_name: str) -> str:
    normalized = normalize_value(value)
    if field_name in {"broad", "strict"} and normalized in {"0", "2"}:
        return "0_or_2"
    return normalized


def looks_like_text(value: Any) -> bool:
    if is_empty(value):
        return False
    text = normalize_value(value)
    if text.lower() in {"true", "false"}:
        return False
    if text.isdigit():
        return False
    return len(text) >= 2


def row_key(values: list[Any], id_index: int | None, text_index: int | None) -> str | None:
    if id_index is not None and id_index < len(values):
        raw_id = normalize_value(values[id_index])
        if raw_id:
            return f"id:{raw_id}"
    if text_index is not None and text_index < len(values):
        raw_text = normalize_value(values[text_index])
        if raw_text:
            return f"text:{raw_text}"
    return None


def find_first_index(headers: list[Any], target: str) -> int | None:
    norm = target.lower()
    for idx, value in enumerate(headers):
        if normalize_header(value) == norm:
            return idx
    return None


def find_nth_index(headers: list[Any], target: str, ordinal: int) -> int | None:
    seen = 0
    norm = target.lower()
    for idx, value in enumerate(headers):
        if normalize_header(value) == norm:
            seen += 1
            if seen == ordinal:
                return idx
    return None


def should_replace(existing: Any, incoming: Any, field_name: str) -> bool:
    if is_empty(existing):
        return True
    if comparable_value(existing, field_name) == comparable_value(incoming, field_name):
        return False
    if field_name in {"cleaned_text", "keyword"} and not looks_like_text(existing) and looks_like_text(incoming):
        return True
    return False


def looks_like_header_row(values: list[Any]) -> bool:
    normalized = {normalize_header(value) for value in values if not is_empty(value)}
    expected = {"id", "cleaned_text", "keyword", "broad", "strict"}
    return len(normalized & expected) >= 3


@dataclass
class WorkbookView:
    path: Path
    sheet_name: str
    rows: list[list[Any]]
    id_index: int | None
    text_index: int | None
    field_indexes: dict[str, int]
    data_start_index: int
    header_mode: str


def detect_layout(rows: list[list[Any]]) -> tuple[dict[str, int], int, str]:
    row1 = rows[0] if len(rows) >= 1 else []
    row2 = rows[1] if len(rows) >= 2 else []

    if looks_like_header_row(row1):
        field_indexes = {
            "id": find_first_index(row1, "id"),
            "cleaned_text": find_first_index(row1, "cleaned_text"),
            "keyword": find_first_index(row1, "keyword"),
            "broad": find_first_index(row1, "broad"),
            "strict": find_first_index(row1, "strict"),
        }
        return ({k: v for k, v in field_indexes.items() if v is not None}, 1, "single_header")

    if looks_like_header_row(row2):
        field_indexes = {
            "id": find_first_index(row2, "id"),
            "cleaned_text": find_first_index(row2, "cleaned_text"),
            "keyword": find_first_index(row2, "keyword"),
            "broad": find_first_index(row2, "broad"),
            "strict": find_first_index(row2, "strict"),
        }
        return ({k: v for k, v in field_indexes.items() if v is not None}, 2, "double_header")

    return (NO_HEADER_INDEXES.copy(), 0, "no_header")


def load_workbook_view(path: Path) -> WorkbookView:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    field_indexes, data_start_index, header_mode = detect_layout(rows)

    return WorkbookView(
        path=path,
        sheet_name=ws.title,
        rows=rows,
        id_index=field_indexes.get("id"),
        text_index=field_indexes.get("cleaned_text"),
        field_indexes=field_indexes,
        data_start_index=data_start_index,
        header_mode=header_mode,
    )


def merge_workbooks(primary_path: Path, input_paths: list[Path], output_path: Path) -> dict[str, Any]:
    primary_wb = load_workbook(primary_path)
    primary_ws = primary_wb[primary_wb.sheetnames[0]]
    primary_rows = [list(row) for row in primary_ws.iter_rows(values_only=True)]
    if len(primary_rows) < 2:
        raise ValueError(f"{primary_path} does not have the expected two-row header")

    primary_header2 = primary_rows[1]
    primary_id_index = find_first_index(primary_header2, "id")
    primary_text_index = find_first_index(primary_header2, "cleaned_text")
    if primary_id_index is None and primary_text_index is None:
        raise ValueError(f"{primary_path} does not have id or cleaned_text columns in row 2")

    primary_keyword_index = find_nth_index(primary_header2, "keyword", 1)
    primary_field_indexes = {
        "id": primary_id_index,
        "cleaned_text": primary_text_index,
        "keyword": primary_keyword_index,
        "broad": find_first_index(primary_header2, "broad"),
        "strict": find_first_index(primary_header2, "strict"),
    }

    row_by_key: dict[str, int] = {}
    for row_idx in range(3, primary_ws.max_row + 1):
        values = [primary_ws.cell(row=row_idx, column=col).value for col in range(1, primary_ws.max_column + 1)]
        key = row_key(values, primary_id_index, primary_text_index)
        if key:
            row_by_key[key] = row_idx

    conflicts: list[dict[str, Any]] = []
    updates = 0
    appended = 0
    source_summaries: list[dict[str, Any]] = []

    for path in input_paths:
        if path == primary_path:
            continue
        view = load_workbook_view(path)
        source_summary = {
            "source_file": path.name,
            "header_mode": view.header_mode,
            "rows_seen": 0,
            "matched_rows": 0,
            "appended_rows": 0,
            "updated_cells": 0,
            "conflicts": 0,
        }

        for values in view.rows[view.data_start_index:]:
            key = row_key(values, view.id_index, view.text_index)
            if key is None:
                continue
            source_summary["rows_seen"] += 1

            target_row = row_by_key.get(key)
            if target_row is None:
                target_row = primary_ws.max_row + 1
                primary_ws.append([None] * max(primary_ws.max_column, PRIMARY_WIDTH))
                row_by_key[key] = target_row
                appended += 1
                source_summary["appended_rows"] += 1
            else:
                source_summary["matched_rows"] += 1

            for field_name, source_index in view.field_indexes.items():
                target_index = primary_field_indexes.get(field_name)
                if target_index is None:
                    target_index = PRIMARY_FILL_MAP.get(field_name)
                if target_index is None or source_index >= len(values):
                    continue

                source_value = values[source_index]
                if is_empty(source_value):
                    continue

                target_cell = primary_ws.cell(row=target_row, column=target_index + 1)
                current_value = target_cell.value

                if should_replace(current_value, source_value, field_name):
                    target_cell.value = source_value
                    updates += 1
                    source_summary["updated_cells"] += 1
                elif comparable_value(current_value, field_name) != comparable_value(source_value, field_name):
                    conflicts.append(
                        {
                            "key": key,
                            "field": field_name,
                            "kept": current_value,
                            "incoming": source_value,
                            "source_file": path.name,
                        }
                    )
                    source_summary["conflicts"] += 1

        source_summaries.append(source_summary)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    primary_wb.save(output_path)

    return {
        "output": str(output_path),
        "primary": str(primary_path),
        "inputs": [str(p) for p in input_paths],
        "rows": primary_ws.max_row,
        "updates": updates,
        "appended_rows": appended,
        "conflicts": len(conflicts),
        "conflict_examples": conflicts[:20],
        "conflict_details": conflicts,
        "sources": source_summaries,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge split annotation xlsx files into one workbook.")
    parser.add_argument("inputs", nargs="+", help="Input xlsx files. The first file is used as the primary template.")
    parser.add_argument("-o", "--output", required=True, help="Output xlsx path.")
    parser.add_argument("--report-json", help="Optional path to write the full merge report as JSON.")
    args = parser.parse_args()

    input_paths = [Path(item).expanduser().resolve() for item in args.inputs]
    output_path = Path(args.output).expanduser().resolve()
    report_path = Path(args.report_json).expanduser().resolve() if args.report_json else None

    result = merge_workbooks(primary_path=input_paths[0], input_paths=input_paths, output_path=output_path)
    console_result = dict(result)
    console_result["conflict_examples"] = result["conflict_details"][:20]
    report_result = dict(result)
    report_result.pop("conflict_examples", None)
    if report_path is not None:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report_result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(console_result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
